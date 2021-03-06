# Importing Libraries
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

def scraping_data():
    # Loading Workbook and selecting colummn and looping through urls
    wb = load_workbook("input.xlsx")
    ws = wb.active
    column_links = ws['C']
    scraped_data = []
    driver = webdriver.Chrome() 
    for url in column_links[1:]:
        link = url.value    
        driver.get(link)
        #driver.maximize_window()
        time.sleep(2)
        title = driver.title 
        meta_description = driver.find_element(By.XPATH, value="html/head/meta[4]").get_attribute("content") 
        meta_keywords = driver.find_element(By.XPATH, value="/html/head/meta[5]").get_attribute("content") 
        canonical_url = driver.find_element(By.XPATH, value="html/head/link[1]").get_attribute("href")
        try:
            about_product = driver.find_element(By.ID, "about_0").text        
        except:
            about_product = "No Data Found"
        try:
            composition = driver.find_element(By.ID, "about_1").text        
        except:
            composition = "No Data Found"
        try:
            features = driver.find_element(By.ID, "about_2").text        
        except:
            features = "No Data Found"
        try:
            how_to_use = driver.find_element(By.ID, "about_3").text        
        except:
            how_to_use = "No Data Found"
        try:
            other_product_info = driver.find_element(By.ID, "about_4").text        
        except:
            other_product_info = "No Data Found"
        scraped_data.append([title, meta_description, meta_keywords, canonical_url, about_product, composition, features, how_to_use, other_product_info])
        time.sleep(2)
        print(title)
    driver.quit()
    
    # Saved to excel
    wb = openpyxl.Workbook()
    ws1 = wb.active
    headers = ["Title", "Meta Description", "Meta Keywords", "Canonical URL", "About Products", "Composition", "Features", "How To Use", "Other Product Info"]    
    ws1.append(headers)
    for row in scraped_data:        
        ws1.append(row)
    wb.save("output_all.xlsx")
scraping_data()